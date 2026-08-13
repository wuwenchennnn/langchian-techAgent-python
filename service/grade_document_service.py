from typing import List, Optional

import io
import logging
import PyPDF2
import openpyxl
import xlrd

from rag import TextSplitter, GradeTextSplitter, Retriever, HybridRetriever
from rag.reranker import create_reranker
from repository.redis_grade_document_store import RedisGradeDocumentStore, bucket_key
from schemas.analysis import StudentScore
from service.bucket_analyzer import BucketAnalyzer
from service.grade_analyzer import GradeAnalyzer
from service.grade_scope_analyzer import GradeScopeAnalyzer

logger = logging.getLogger(__name__)


class GradeDocumentService:
    """成绩文档服务：负责 PDF/Excel 解析、按（年级+班级）桶存储、RAG 流程编排与多考试分析"""

    PDF_MAGIC = b"%PDF-"
    XLSX_MAGIC = b"PK\x03\x04"

    SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}

    def __init__(self):
        self.document_store = RedisGradeDocumentStore()
        self.splitter = TextSplitter()
        self.grade_splitter = GradeTextSplitter()
        self.vector_retriever = Retriever()
        self.reranker, self._reranker_type = create_reranker()
        logger.info("[重排序器] 当前使用: %s", self._reranker_type)
        # 内存缓存按桶（bucket_id）隔离
        self._bucket_analyzers: dict[str, BucketAnalyzer] = {}
        self._hybrid_retrievers: dict[str, HybridRetriever] = {}
        # 年级级缓存（跨班对比）
        self._grade_analyzers: dict[str, GradeScopeAnalyzer] = {}
        self._grade_hybrids: dict[str, HybridRetriever] = {}

    # ---------- 上传与存储 ----------
    def upload_and_store(self, grade: str, class_name: str, exam_name: str, file) -> str:
        """解析并存储一份考试文档到（年级+班级）桶，随后重建桶内分析器与检索索引"""
        if not file:
            raise ValueError("文件不能为空")

        filename = getattr(file, "filename", "")
        if filename:
            ext = filename.lower()
            dot_idx = ext.rfind(".")
            ext = ext[dot_idx:] if dot_idx != -1 else ""
            if ext not in self.SUPPORTED_EXTENSIONS:
                raise ValueError(f"仅支持 PDF / Excel（.xlsx / .xls）文件，当前文件: {filename}")

        file.file.seek(0)
        file_content = file.file.read()

        if not file_content:
            raise ValueError("文件内容为空")

        if file_content.startswith(self.PDF_MAGIC):
            extracted_text = self._extract_text_from_pdf(file_content)
        elif file_content.startswith(self.XLSX_MAGIC):
            extracted_text = self._extract_text_from_xlsx(file_content)
        else:
            ext = filename.lower() if filename else ""
            dot_idx = ext.rfind(".")
            ext = ext[dot_idx:] if dot_idx != -1 else ""
            if ext == ".xls":
                extracted_text = self._extract_text_from_xls(file_content)
            else:
                raise ValueError("无法识别的文件格式，请上传 PDF 或 Excel 文件")

        # 结构化分析
        analyzer = GradeAnalyzer()
        records = analyzer.parse(extracted_text)
        logger.info(
            "解析到 %d 条成绩记录，%d 名学生，%d 门科目",
            len(records), len(analyzer._student_names), len(analyzer._subjects)
        )

        # 结构化语义分块，优先使用记录分块，回退到固定分块
        if records and analyzer._student_names:
            chunks = self.grade_splitter.split_by_records(
                records, analyzer._student_names, analyzer._subjects
            )
        else:
            chunks = self.splitter.split(extracted_text)

        # 每个 chunk 前置「年级班级·考试名」元数据行，保证召回片段自带考试归属
        tagged_chunks = [
            f"【{grade}{class_name}·{exam_name}】\n{chunk}"
            for chunk in chunks
        ]

        chunk_records: List[dict] = []
        if tagged_chunks:
            try:
                vectors = self.vector_retriever.embed_documents(tagged_chunks)
                chunk_records = [
                    {
                        "index": index,
                        "content": chunk,
                        "embedding": vector,
                        "grade": grade,
                        "className": class_name,
                        "examName": exam_name,
                    }
                    for index, (chunk, vector) in enumerate(zip(tagged_chunks, vectors))
                ]
            except Exception as e:
                logger.warning("Embedding 失败（将退回为全文检索）: %s", e)

        serialized_records = [
            {"student_name": r.student_name, "subject": r.subject, "score": r.score}
            for r in records
        ]
        self.document_store.store_document(
            grade, class_name, exam_name,
            text=extracted_text,
            records=serialized_records,
            chunks=chunk_records,
            filename=filename,
            student_count=len(analyzer._student_names),
            subject_count=len(analyzer._subjects),
        )

        self._rebuild_bucket_state(grade, class_name)
        return extracted_text

    # ---------- 桶状态（内存缓存 + Redis 重建） ----------
    def _rebuild_bucket_state(self, grade: str, class_name: str):
        """从 Redis 重建桶分析器与混合检索引擎（上传/删除后或缓存缺失时调用）"""
        bkey = bucket_key(grade, class_name)
        metadata = self.document_store.get_bucket_metadata(grade, class_name)
        if metadata and metadata.get("docs"):
            analyzers = {}
            exam_order = []
            for doc in metadata["docs"]:
                exam_name = doc.get("examName", "")
                raw_records = self.document_store.get_records(grade, class_name, exam_name)
                if not raw_records:
                    continue
                records = [
                    StudentScore(
                        student_name=r.get("student_name", ""),
                        subject=r.get("subject", ""),
                        score=float(r.get("score", 0)),
                    )
                    for r in raw_records
                ]
                analyzers[exam_name] = GradeAnalyzer.from_records(records)
                exam_order.append(exam_name)
            self._bucket_analyzers[bkey] = BucketAnalyzer(
                grade, class_name, analyzers, exam_order
            )
        else:
            self._bucket_analyzers.pop(bkey, None)

        chunks = self.document_store.get_bucket_chunks(grade, class_name)
        if chunks:
            hybrid = HybridRetriever(self.vector_retriever)
            hybrid.index_chunks(chunks)
            self._hybrid_retrievers[bkey] = hybrid
        else:
            self._hybrid_retrievers.pop(bkey, None)
        self._invalidate_grade(grade)

    def get_analyzer(self, grade: str, class_name: Optional[str] = None):
        """获取分析器：指定班级返回桶分析器；不指定返回年级级分析器（跨班对比）"""
        if class_name:
            bkey = bucket_key(grade, class_name)
            analyzer = self._bucket_analyzers.get(bkey)
            if analyzer is None:
                metadata = self.document_store.get_bucket_metadata(grade, class_name)
                if not metadata or not metadata.get("docs"):
                    return None
                self._rebuild_bucket_state(grade, class_name)
                analyzer = self._bucket_analyzers.get(bkey)
            return analyzer

        grade_analyzer = self._grade_analyzers.get(grade)
        if grade_analyzer is None:
            self._rebuild_grade_state(grade)
            grade_analyzer = self._grade_analyzers.get(grade)
        return grade_analyzer

    # ---------- 年级级状态（跨班对比） ----------
    def _invalidate_grade(self, grade: str):
        """班级数据变化时失效对应年级的缓存"""
        self._grade_analyzers.pop(grade, None)
        self._grade_hybrids.pop(grade, None)

    def _collect_grade_chunks(self, grade: str) -> List[dict]:
        """收集该年级所有班级桶的聚合 chunks"""
        all_chunks = []
        for meta in self.document_store.list_buckets():
            if meta.get("grade") != grade:
                continue
            class_name = meta.get("className", "")
            if not class_name:
                continue
            chunks = self.document_store.get_bucket_chunks(grade, class_name)
            if chunks:
                all_chunks.extend(chunks)
        return all_chunks

    def _get_grade_full_text(self, grade: str) -> Optional[str]:
        """年级级兜底：无向量时拼接该年级全部桶原文"""
        parts = []
        for meta in self.document_store.list_buckets():
            if meta.get("grade") != grade:
                continue
            class_name = meta.get("className", "")
            if not class_name:
                continue
            text = self.document_store.get_bucket_full_text(grade, class_name)
            if text:
                parts.append(f"【{grade}{class_name}】\n{text}")
        return "\n\n---\n\n".join(parts) if parts else None

    def _rebuild_grade_state(self, grade: str):
        """从 Redis 重建年级级分析器与检索索引（跨班对比用）"""
        buckets = {}
        for meta in self.document_store.list_buckets():
            if meta.get("grade") != grade:
                continue
            class_name = meta.get("className", "")
            if not class_name:
                continue
            bkey = bucket_key(grade, class_name)
            bucket_analyzer = self._bucket_analyzers.get(bkey)
            if bucket_analyzer is None:
                self._rebuild_bucket_state(grade, class_name)
                bucket_analyzer = self._bucket_analyzers.get(bkey)
            if bucket_analyzer is not None:
                buckets[class_name] = bucket_analyzer

        if buckets:
            self._grade_analyzers[grade] = GradeScopeAnalyzer(grade, buckets)
        else:
            self._grade_analyzers.pop(grade, None)

        chunks = self._collect_grade_chunks(grade)
        if chunks:
            hybrid = HybridRetriever(self.vector_retriever)
            hybrid.index_chunks(chunks)
            self._grade_hybrids[grade] = hybrid
        else:
            self._grade_hybrids.pop(grade, None)

    def _get_hybrid(self, grade: str, class_name: str, chunks: List[dict]) -> Optional[HybridRetriever]:
        bkey = bucket_key(grade, class_name)
        hybrid = self._hybrid_retrievers.get(bkey)
        if hybrid is None and chunks:
            hybrid = HybridRetriever(self.vector_retriever)
            hybrid.index_chunks(chunks)
            self._hybrid_retrievers[bkey] = hybrid
        return hybrid

    def get_relevant_content(self, grade: str, class_name: Optional[str], message: str) -> Optional[str]:
        """召回相关片段：单班按桶检索；年级级（class_name 为空）合并该年级全部班级检索"""
        if class_name:
            chunks = self.document_store.get_bucket_chunks(grade, class_name)
            if not chunks:
                return self.document_store.get_bucket_full_text(grade, class_name)

            hybrid = self._get_hybrid(grade, class_name, chunks)
            if hybrid:
                try:
                    candidates, scored = hybrid.retrieve(
                        message,
                        top_k=None,
                        candidate_multiplier=3,
                    )
                    if scored:
                        reranked = self.reranker.rerank(message, scored)
                        contents = [c.get("content", "") for c in reranked]
                        if contents:
                            logger.info(
                                "[混合检索+重排序] bucket=%s | 候选=%d → 返回=%d | 引擎=%s",
                                bucket_key(grade, class_name), len(scored), len(contents),
                                self._reranker_type
                            )
                            return "\n\n---\n\n".join(contents)
                except Exception as e:
                    logger.warning("混合检索失败，回退到纯向量检索: %s", e)

            # 兜底：纯向量检索
            return self.vector_retriever.retrieve(message, chunks)

        # 年级级检索：合并该年级全部班级 chunks
        chunks = self._collect_grade_chunks(grade)
        if not chunks:
            return self._get_grade_full_text(grade)

        hybrid = self._grade_hybrids.get(grade)
        if hybrid is None:
            hybrid = HybridRetriever(self.vector_retriever)
            hybrid.index_chunks(chunks)
            self._grade_hybrids[grade] = hybrid
        if hybrid:
            try:
                candidates, scored = hybrid.retrieve(
                    message,
                    top_k=None,
                    candidate_multiplier=3,
                )
                if scored:
                    reranked = self.reranker.rerank(message, scored)
                    contents = [c.get("content", "") for c in reranked]
                    if contents:
                        logger.info(
                            "[年级级混合检索+重排序] grade=%s | 候选=%d → 返回=%d | 引擎=%s",
                            grade, len(scored), len(contents), self._reranker_type
                        )
                        return "\n\n---\n\n".join(contents)
            except Exception as e:
                logger.warning("年级级混合检索失败，回退到纯向量检索: %s", e)

        return self.vector_retriever.retrieve(message, chunks)

    # ---------- 桶列表与删除 ----------
    def list_buckets(self) -> List[dict]:
        return self.document_store.list_buckets()

    def delete_document(self, grade: str, class_name: str, exam_name: str):
        """删除桶内单份考试文档并重建桶状态"""
        self.document_store.delete_document(grade, class_name, exam_name)
        self._rebuild_bucket_state(grade, class_name)

    def delete_bucket(self, grade: str, class_name: str):
        """删除整个桶（文档 + 聚合检索源 + 内存缓存）"""
        self.document_store.delete_bucket(grade, class_name)
        bkey = bucket_key(grade, class_name)
        self._bucket_analyzers.pop(bkey, None)
        self._hybrid_retrievers.pop(bkey, None)
        self._invalidate_grade(grade)

    # ---------- PDF / Excel 解析 ----------
    def _extract_text_from_pdf(self, pdf_content: bytes) -> str:
        text = ""
        try:
            reader = PyPDF2.PdfReader(io.BytesIO(pdf_content))
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                page_text = page.extract_text() or ""
                text += page_text + "\n"
        except Exception as e:
            raise ValueError(f"PDF解析失败: {str(e)}")
        return text.strip()

    def _extract_text_from_xlsx(self, xlsx_content: bytes) -> str:
        text_parts = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_content), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"【工作表: {sheet_name}】")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(
                        str(cell) if cell is not None else ""
                        for cell in row
                    )
                    if row_text.strip():
                        text_parts.append(row_text)
            wb.close()
        except Exception as e:
            raise ValueError(f"Excel(.xlsx)解析失败: {str(e)}")
        return "\n".join(text_parts).strip()

    def _extract_text_from_xls(self, xls_content: bytes) -> str:
        text_parts = []
        try:
            wb = xlrd.open_workbook(file_contents=xls_content)
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                text_parts.append(f"【工作表: {ws.name}】")
                for row_idx in range(ws.nrows):
                    row_values = ws.row_values(row_idx)
                    row_text = "\t".join(
                        str(cell) if cell != "" else ""
                        for cell in row_values
                    )
                    if row_text.strip():
                        text_parts.append(row_text)
        except Exception as e:
            raise ValueError(f"Excel(.xls)解析失败: {str(e)}")
        return "\n".join(text_parts).strip()
